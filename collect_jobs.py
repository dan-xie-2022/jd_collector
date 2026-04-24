'''
LinkedIn Job Collector
- Searches LinkedIn for jobs (no auto-apply)
- Scores each job against your resume using Claude
- Outputs a sorted Excel spreadsheet
'''

import time
import os
import re
import json
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from openai import OpenAI
import pandas as pd
from dotenv import load_dotenv
from utils import create_driver, login, get_job_description

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

# ─── CONFIG ───────────────────────────────────────────────────────────────────

LINKEDIN_EMAIL    = os.environ["LINKEDIN_EMAIL"]
LINKEDIN_PASSWORD = os.environ["LINKEDIN_PASSWORD"]
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")

ENABLE_SCORING = True          # True = use Claude API to score jobs, False = skip scoring (no API cost)

# ─── MODE ────────────────────────────────────────────────────────────────────
# "default" = Solutions Engineer / TAM / Customer Success 等岗位
# "cs"      = Computer Science 相关开发岗位（Software Engineer, Backend, etc.）
MODE = "default"

# ─── SEARCH TERMS BY MODE ────────────────────────────────────────────────────
_SEARCH_TERMS_DEFAULT = [
    # Product Manager
    "AI Product Manager",
    "Product Manager AI",
    "Senior Product Manager GenAI",
    "Product Manager LLM",
    "Product Manager Machine Learning",
    "AI Product Lead",
    # Innovation / Strategy
    "AI Innovation Manager",
    "AI Innovation Lead",
    "AI Strategy",
    "GenAI Consultant",
    "AI Strategist",
    "AI Strategy Lead",
    # Forward Deployment / Implementation
    "AI Forward Deployment",
    "Forward Deployed AI",
    "AI Implementation Manager",
    "AI Deployment Specialist",
]

_SEARCH_TERMS_CS = [
    # Examples — customize these to your target roles
    "Software Engineer",
    "Backend Developer",
    "Data Analyst",
    "Python Developer",
    "DevOps Engineer",
]

SEARCH_TERMS = _SEARCH_TERMS_CS if MODE == "cs" else _SEARCH_TERMS_DEFAULT

# ─── JD RELEVANCE FILTER ─────────────────────────────────────────────────────
# Jobs whose title or JD matches these patterns are auto-excluded
EXCLUDE_INDUSTRIES = re.compile(
    r'pharma|medical|clinical|biotech|nursing|dental|implant|'
    r'semiconductor|NAND|wafer|fab\b|chip design|'
    r'civil engineer|construction|建筑|mechanical engineer|'
    r'chemical engineer|petroleum|oil\s*&?\s*gas|'
    r'food scientist|agronomist|veterinar',
    re.IGNORECASE
)
# Jobs whose title or JD contains these keywords get a relevance boost
_RELEVANT_KEYWORDS_DEFAULT = re.compile(
    r'AI product|GenAI|generative AI|LLM|large language model|'
    r'conversational AI|agentic|RAG|chatbot|NLP|'
    r'product manager|product lead|product owner|'
    r'machine learning|ML platform|AI platform|'
    r'SaaS|B2B|enterprise|scale-up|startup|'
    r'roadmap|discovery|evaluation|observability|'
    r'Langfuse|LangChain|OpenAI|Anthropic|Claude|GPT',
    re.IGNORECASE
)
_RELEVANT_KEYWORDS_CS = re.compile(
    r'Python|SQL|pandas|data.?analysis|ETL|'
    r'database|MySQL|PostgreSQL|SQLite|'
    r'BI|Tableau|Power\s*BI|Excel|'
    r'automation|scripting|report|dashboard|'
    r'junior|entry.?level|graduate|'
    r'computer science',
    re.IGNORECASE
)
RELEVANT_KEYWORDS = _RELEVANT_KEYWORDS_CS if MODE == "cs" else _RELEVANT_KEYWORDS_DEFAULT

SEARCH_LOCATION = "London"
DATE_FILTER     = "Past 24 hours"      # "Past 24 hours", "Past week", "Past month"
MAX_JOBS_PER_SEARCH = 5        # max jobs to collect per search term

RESUME_SUMMARY = """
Name: Dan Xie
Current Role: AI Product Manager at Ohme Energy (EV Charging Technology), London, UK — July 2025 to Present
Education: PhD in Marketing, HEC Paris (2018–2024, Elyette Roux Prize for Best PhD Thesis); MSc in Economics & Psychology, Panthéon-Sorbonne (2016–2018, International Excellence Scholarship); BSc in Applied Psychology, Shaanxi Normal University (2009–2013, Best of Class Dissertation Award)
Languages: English (fluent), Chinese (native), French (conversational)
Tools: Python, SQL, R, Snowflake
Conversational AI skills: Multi-agent architecture, RAG design, LLM orchestration, tool-use agents, human-in-the-loop (HITL), agentic workflow design, context window management
Prompt Engineering: System prompt design, multi-layer prompt architecture, guardrail calibration
AI Evaluation: Golden dataset design, LLM-as-judge, RAG retrieval evaluation, regression testing; RAGAS-based eval frameworks; live trace monitoring in Langfuse
QA: Expertise in QA testing agentic products
Other PM skills: PRD authoring, agile backlog management, product discovery, metrics framework design, staged rollout, shadow testing, user research, cross-functional stakeholder management

Experience:
- Ohme Energy, AI Product Manager (Jul 2025–Present, London):
  • Led end-to-end product ownership of Otis — conversational AI chatbot serving 400K+ users; 40% automated resolution at 10% of the cost of Intercom Fin AI
  • Product discovery via K-means clustering on 30,000+ customer queries; designed full metrics framework (conversations, resolutions, escalations, cost per resolution) with live observability in Langfuse
  • Designed human-in-the-loop interrupt handling for multi-step agentic workflows (escalation trigger logic, state persistence, graceful failure modes)
  • Shipped CC Copilot (action-taking agent, NPS 10/10 from early adopters) and Troubleshooter Email (30% auto-resolution of inbound support emails)
  • Owned AI eval strategy: golden datasets from 30K+ real queries, RAGAS-based eval, live Langfuse trace monitoring
  • Two ML features: computer vision pipeline for EV charger photo classification (−16.7% onboarding time); LightGBM charger state model (+67% feature adoption, −17% undercharging)
  • Co-owned company-wide AI agent platform roadmap: MCP server, multi-agent memory, agent front-ends
  • Reported directly to CEO; cross-functional delivery across Engineering, Customer Service, Data Engineering, Product leadership

- Finres, Product Manager (May 2024–Jun 2025, London/Paris) — VC-backed Climate Fintech:
  • Launched climate risk scoring and visualisation features for portfolio stress testing; adopted by 2 French financial sector clients
  • Introduced Agile workflows and R&D KPIs; reduced experimentation cycles by 25%
  • LLM-powered user segmentation → 30% adoption uplift for key product features
  • Owned delivery in PREVENT — a €1M+ EU consortium across 20+ partners and 8 countries

- HEC Paris, PhD Researcher in Marketing (Sep 2018–May 2024, Paris):
  • Quantitative research: propensity score matching, instrumental variables, difference-in-differences
  • Data science: cluster analysis, NLP, predictive ML models (Python/R) on large datasets
  • Managed 15,000+ participants across survey studies; designed 30+ psychometric scales and experiments
  • Lectured econometrics and statistics to Master students; improved course evaluation by 15%

- University of Liverpool Management School, Guest Lecturer — AI in Business (Mar 2026)
- Shaanxi Normal University, Research Manager (2013–2016, Xi'an): built school innovation index for a district of 600K residents

Target roles: AI Product Manager, Senior Product Manager (AI/GenAI/LLM), Head of AI Product, Product Lead (AI)
Target industry: Technology, AI/GenAI, SaaS, FinTech, Climate Tech, Enterprise Software
Preferred company type: Scale-ups or established tech companies actively building AI/GenAI products, strong engineering culture
Location: London, UK (open to hybrid or remote; not looking to relocate)
Seniority: Mid-to-senior individual contributor (not seeking director/VP/C-suite)
"""

LIST_DIR      = os.path.join(os.path.dirname(os.path.abspath(__file__)), "List")
OUTPUT_FILE   = f"job_results_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
SEEN_IDS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "seen_jobs.json")

SEEN_EXPIRE_DAYS = 30  # seen jobs older than this are forgotten

# Load overrides from webapp_config.json if it exists (set by the Streamlit web app)
_WEBAPP_CONFIG = os.path.join(os.path.dirname(os.path.abspath(__file__)), "webapp_config.json")
if os.path.exists(_WEBAPP_CONFIG):
    with open(_WEBAPP_CONFIG, encoding="utf-8") as _f:
        _wc = json.load(_f)
    SEARCH_TERMS        = _wc.get("search_terms", SEARCH_TERMS)
    SEARCH_LOCATION     = _wc.get("search_location", SEARCH_LOCATION)
    DATE_FILTER         = _wc.get("date_filter", DATE_FILTER)
    MAX_JOBS_PER_SEARCH = _wc.get("max_jobs_per_search", MAX_JOBS_PER_SEARCH)
    RESUME_SUMMARY      = _wc.get("resume_summary", RESUME_SUMMARY)
    ENABLE_SCORING      = _wc.get("enable_scoring", ENABLE_SCORING)

def load_seen_ids():
    if os.path.exists(SEEN_IDS_FILE):
        with open(SEEN_IDS_FILE) as f:
            data = json.load(f)
        # Migrate from old list format to {id: timestamp} dict
        if isinstance(data, list):
            now = datetime.now().isoformat()
            return {jid: now for jid in data}
        # Expire old entries
        cutoff = datetime.now().timestamp() - SEEN_EXPIRE_DAYS * 86400
        return {jid: ts for jid, ts in data.items()
                if datetime.fromisoformat(ts).timestamp() > cutoff}
    return {}

def save_seen_ids(seen_ids):
    with open(SEEN_IDS_FILE, "w") as f:
        json.dump(seen_ids, f)

# ─── JOB SEARCH ───────────────────────────────────────────────────────────────

DATE_MAP = {
    "Past 24 hours": "r86400",
    "Past week":     "r604800",
    "Past month":    "r2592000",
}

# LinkedIn geoIds for location filtering
GEO_IDS = {
    "China":          "102890883",
    "Shanghai":       "102772228",
    "Shenzhen":       "101591017",
    "Beijing":        "101780494",
    "United Kingdom": "101165590",
    "London":         "90009496",
}

def build_search_url(search_term, location, date_filter, remote_only=False):
    from urllib.parse import quote
    term_enc = quote(search_term)
    loc_enc  = quote(location)
    date_param = DATE_MAP.get(date_filter, "r604800")
    geo_id = GEO_IDS.get(location, "")
    url = (
        f"https://www.linkedin.com/jobs/search/"
        f"?keywords={term_enc}&location={loc_enc}"
        f"&f_TPR={date_param}&sortBy=DD"
    )
    if geo_id:
        url += f"&geoId={geo_id}"
    if remote_only:
        url += "&f_WT=2"  # LinkedIn filter: Remote only
    return url

def get_job_cards(driver, max_jobs):
    jobs = []
    seen_ids = set()

    # wait for results to load
    time.sleep(3)

    for scroll_attempt in range(15):
        # Try multiple selectors for job cards
        cards = driver.find_elements(By.CSS_SELECTOR, "li[data-occludable-job-id]")
        if not cards:
            cards = driver.find_elements(By.CSS_SELECTOR, ".scaffold-layout__list-container li")
        if not cards:
            cards = driver.find_elements(By.CSS_SELECTOR, ".jobs-search-results__list li")

        for card in cards:
            try:
                job_id = card.get_attribute("data-occludable-job-id") or card.get_attribute("data-job-id") or card.get_attribute("data-entity-urn")
                if not job_id or job_id in seen_ids:
                    continue
                seen_ids.add(job_id)

                # Title and URL
                title_el = None
                for sel in ["a.job-card-list__title--link", "a.job-card-container__link", "a[href*='/jobs/view/']", ".job-card-list__title"]:
                    try:
                        title_el = card.find_element(By.CSS_SELECTOR, sel)
                        break
                    except NoSuchElementException:
                        continue
                if not title_el:
                    continue
                title = title_el.text.strip()
                url = title_el.get_attribute("href") or ""
                if "?" in url:
                    url = url.split("?")[0]

                # Company
                company = ""
                for sel in [".job-card-container__primary-description", ".artdeco-entity-lockup__subtitle span", ".job-card-container__company-name"]:
                    try:
                        company = card.find_element(By.CSS_SELECTOR, sel).text.strip()
                        if company:
                            break
                    except NoSuchElementException:
                        continue

                # Location
                location = ""
                for sel in [".job-card-container__metadata-item", ".artdeco-entity-lockup__caption li", ".job-card-container__metadata-wrapper li"]:
                    try:
                        location = card.find_element(By.CSS_SELECTOR, sel).text.strip()
                        if location:
                            break
                    except NoSuchElementException:
                        continue

                # Date
                date_text = ""
                for sel in ["time", ".job-card-container__listed-time", ".job-card-list__footer-wrapper time"]:
                    try:
                        el = card.find_element(By.CSS_SELECTOR, sel)
                        date_text = el.get_attribute("datetime") or el.text.strip()
                        if date_text:
                            break
                    except NoSuchElementException:
                        continue

                # Easy Apply badge
                easy_apply = False
                try:
                    badge_text = card.text.lower()
                    easy_apply = "easy apply" in badge_text or "快速申请" in badge_text
                except Exception:
                    pass

                if not title:
                    continue

                jobs.append({
                    "job_id":      job_id,
                    "title":       title,
                    "company":     company,
                    "location":    location,
                    "date_text":   date_text,
                    "url":         url,
                    "easy_apply":  easy_apply,
                    "description": "",
                })

                if len(jobs) >= max_jobs:
                    return jobs
            except Exception:
                continue

        # Scroll down in the results panel
        try:
            scroll_container = driver.find_element(By.CSS_SELECTOR, ".scaffold-layout__list-container, .jobs-search-results-list, .jobs-search__results-list")
            driver.execute_script("arguments[0].scrollTop += 800", scroll_container)
        except (NoSuchElementException, Exception):
            try:
                driver.execute_script("window.scrollBy(0, 800);")
            except Exception:
                pass
        time.sleep(1.5)

    return jobs

# ─── DEEPSEEK SCORING ─────────────────────────────────────────────────────────

_client = None

def _get_client():
    global _client
    if _client is None:
        _client = OpenAI(api_key=DEEPSEEK_API_KEY, base_url="https://api.deepseek.com")
    return _client

_SCORE_SYSTEM = f"""You are a strict career advisor. Score how well a job matches the candidate.

SCORING RULES (follow strictly):
- 8–10: Perfect match — AI/GenAI PM role (Product Manager, AI PM, or AI Product Owner) + tech/SaaS/fintech/climate-tech industry + mid-to-senior IC level + London or UK Remote
- 6–7: Strong match — most criteria met with minor gaps (e.g. "Senior PM" at an AI-adjacent company, hybrid role outside London, or AI PM role in a slightly adjacent industry like insurtech or energy tech)
- 4–5: Partial match — some overlap but meaningful gaps (e.g. PM role with AI exposure but not AI-first, wrong seniority, or B2B SaaS without AI focus)
- 1–3: Poor match — fundamentally misaligned (sales, engineering, ops, or non-PM roles) or unrelated industry

Hard caps:
- Director/VP/C-suite roles: cap at 3
- Roles requiring no AI/ML experience: cap at 4
- Roles in unrelated industries (manufacturing, healthcare ops, construction, retail ops): cap at 3
- Roles outside the UK (unless fully remote and explicitly open to UK): cap at 3

Action thresholds:
- Below 3 — Flag as poor fit, do not apply
- 3–6 — Apply with Commercial CV as-is, no tailoring needed
- Above 6 — Provide full tailoring suggestions

CANDIDATE PROFILE:
{RESUME_SUMMARY}

Respond in this exact format (nothing else):
SCORE: <1-10>
REASON: <one sentence why>
MATCH_TAGS: <comma-separated keywords that match, e.g. "bilingual, SaaS, technical support">"""


def _build_negative_examples_str() -> str:
    """Load 'not interested' jobs from DB and format as prompt negative examples."""
    try:
        import db as _db
        examples = _db.get_negative_examples(limit=15)
        if not examples:
            return ""
        lines = [
            f'  - "{e["title"]}" @ {e["company"]}'
            + (f': {e["score_reason"]}' if e.get("score_reason") else "")
            for e in examples
        ]
        return (
            "\n\nNEGATIVE EXAMPLES — the candidate marked these as NOT INTERESTED "
            "(score similar jobs LOWER, apply same reasoning to penalise alike roles):\n"
            + "\n".join(lines)
        )
    except Exception:
        return ""


def score_job(title, company, description):
    try:
        neg_str = _build_negative_examples_str()
        msg = _get_client().chat.completions.create(
            model="deepseek-chat",
            max_tokens=200,
            messages=[
                {"role": "system", "content": _SCORE_SYSTEM},
                {"role": "user", "content": (
                    f"JOB:\nTitle: {title}\nCompany: {company}\n"
                    f"Description (excerpt):\n{description[:8000]}"
                    f"{neg_str}"
                )},
            ],
        )
        text = msg.choices[0].message.content.strip()
        score_match = re.search(r"SCORE:\s*(\d+)", text)
        reason_match = re.search(r"REASON:\s*(.+)", text)
        tags_match = re.search(r"MATCH_TAGS:\s*(.+)", text)
        score  = int(score_match.group(1)) if score_match else 5
        reason = reason_match.group(1).strip() if reason_match else ""
        tags   = tags_match.group(1).strip() if tags_match else ""
        return score, reason, tags
    except Exception as e:
        print(f"  DeepSeek error: {e}")
        return 5, "Could not score", ""

# ─── PRE-FILTER ───────────────────────────────────────────────────────────────

# Keywords that indicate a job is NOT a good match (case-insensitive)
_TITLE_SKIP_DEFAULT = [
    # Pure engineering — not PM roles
    "software engineer", "backend engineer", "frontend engineer",
    "fullstack", "full stack", "full-stack", "ml engineer", "data engineer",
    "cloud engineer", "devops", "platform engineer", "infrastructure",
    "hardware", "firmware", "embedded",
    # Sales / commercial
    "sales manager", "account executive", "account manager", "business development",
    "pre-sales", "solutions engineer", "customer success",
    # Too junior
    "intern", "graduate scheme", "apprentice",
    # Too senior
    "chief product officer", "CPO", "vice president", "VP of product",
    "C-suite", "CTO", "CEO",
    # Unrelated fields
    "clinical trial", "medical", "FAE", "hardware sales",
    "焊接", "模拟IC", "功率器件", "设计院", "汽车零部件",
]
_TITLE_SKIP_CS = [
    "hardware", "clinical trial", "medical solution", "FAE",
    "sales manager", "security sales",
    "焊接", "模拟IC", "功率器件", "设计院", "汽车零部件",
    "director", "senior director", "vice president", "VP ", "head of",
    "principal", "lead architect", "staff engineer",
    "customer success", "account manager", "pre-sales",
    "senior software", "senior developer", "senior engineer",
    "frontend", "full stack", "fullstack", "full-stack",
    "Java", "Go ", "Rust", "C++", "React", "Node",
    "DevOps", "SRE", "infrastructure", "Kubernetes",
    "machine learning", "ml engineer", "blockchain",
]
TITLE_SKIP_KEYWORDS = _TITLE_SKIP_CS if MODE == "cs" else _TITLE_SKIP_DEFAULT

def has_chinese(text):
    return bool(re.search(r'[\u4e00-\u9fff]', text))

_NON_UK_LOCATIONS = re.compile(
    r'United States|USA|\bUS\b|Canada|'
    r'China|Shanghai|Beijing|Shenzhen|'
    r'India|Bangalore|Hyderabad|Mumbai|'
    r'Germany|France|Netherlands|Spain|Italy|Poland|'
    r'Australia|Japan|Singapore|Brazil|'
    r'\bNY\b|\bCA\b|\bTX\b|\bWA\b|\bMA\b|\bIL\b|'
    r'New York|San Francisco|Los Angeles|Seattle|Boston|Chicago|Austin|'
    r'Berlin|Paris|Amsterdam|Madrid|Warsaw|Tokyo|Toronto|Vancouver|Sydney',
    re.IGNORECASE
)

def is_acceptable_location(location):
    """
    Keep: London jobs, UK Remote jobs, unknown location.
    Skip: non-UK jobs, and non-London onsite/hybrid in UK.
    """
    if not location:
        return True  # unknown, keep
    loc = location.strip()
    is_uk = bool(re.search(r'United Kingdom|UK\b|\bEngland\b|London|Manchester|Edinburgh|Birmingham', loc, re.IGNORECASE))
    is_remote = bool(re.search(r'Remote', loc, re.IGNORECASE))
    is_london = bool(re.search(r'London', loc, re.IGNORECASE))
    # Clearly non-UK → skip
    if not is_uk and _NON_UK_LOCATIONS.search(loc):
        return False
    # UK Remote → keep
    if is_uk and is_remote:
        return True
    # London (any work mode) → keep
    if is_london:
        return True
    # Other UK cities, not remote → skip
    if is_uk and not is_remote:
        return False
    return True  # ambiguous, keep

def should_fetch_jd(title, company):
    title_lower = title.lower()
    if has_chinese(title):
        return False
    for kw in TITLE_SKIP_KEYWORDS:
        if kw.lower() in title_lower:
            return False
    return True

# ─── EXCEL OUTPUT ─────────────────────────────────────────────────────────

def _save_excel(jobs, output_path):
    """Save jobs list to Excel (used for incremental saves during scoring)."""
    df = pd.DataFrame(jobs, columns=[
        "relevance_score", "relevance_reason", "match_tags",
        "title", "company", "location", "date_text",
        "easy_apply", "search_term", "url", "description"
    ])
    df.columns = [
        "相关性评分", "评分理由", "匹配关键词",
        "职位名称", "公司", "地点", "发布时间",
        "Easy Apply", "搜索词", "链接", "职位描述"
    ]
    df.to_excel(output_path, index=False, sheet_name="职位列表")

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    driver = create_driver()
    all_jobs = []
    seen_ids = load_seen_ids()
    print(f"Already seen: {len(seen_ids)} jobs (will skip)")

    try:
        login(driver, LINKEDIN_EMAIL, LINKEDIN_PASSWORD)

        # Two search rounds: London (all) + UK (remote only)
        search_rounds = [
            (SEARCH_LOCATION,  False, "London"),
            ("United Kingdom", True,  "UK Remote"),
        ]

        for round_location, round_remote, round_label in search_rounds:
            print(f"\n{'#'*60}")
            print(f"# Round: {round_label}")
            print(f"{'#'*60}")

            for term in SEARCH_TERMS:
                print(f"\n{'='*60}")
                print(f"Searching: {term} ({round_label})")
                try:
                    url = build_search_url(term, round_location, DATE_FILTER, remote_only=round_remote)
                    driver.get(url)
                    time.sleep(3)

                    jobs = get_job_cards(driver, MAX_JOBS_PER_SEARCH)
                    print(f"  Found {len(jobs)} jobs")

                    skipped = 0
                    for i, job in enumerate(jobs):
                        job["search_term"] = f"{term} ({round_label})"
                        if job["job_id"] in seen_ids:
                            skipped += 1
                            continue
                        seen_ids[job["job_id"]] = datetime.now().isoformat()  # mark as seen
                        if not is_acceptable_location(job["location"]):
                            print(f"  [{i+1}/{len(jobs)}] ✗ SKIP (地点) {job['title']} @ {job['company']} ({job['location']})")
                            continue
                        if not should_fetch_jd(job["title"], job["company"]):
                            print(f"  [{i+1}/{len(jobs)}] ✗ SKIP {job['title']} @ {job['company']}")
                            continue
                        # Check title against industry exclude list
                        if EXCLUDE_INDUSTRIES.search(job["title"]):
                            print(f"  [{i+1}/{len(jobs)}] ✗ EXCLUDE (industry) {job['title']} @ {job['company']}")
                            continue
                        job["description"] = get_job_description(driver, job["url"])
                        time.sleep(1)
                        # Check JD against industry exclude list
                        if job["description"] and EXCLUDE_INDUSTRIES.search(job["description"][:500]):
                            print(f"  [{i+1}/{len(jobs)}] ✗ EXCLUDE (JD) {job['title']} @ {job['company']}")
                            continue
                        # Tag relevance
                        jd_text = job["title"] + " " + job.get("description", "")
                        hits = RELEVANT_KEYWORDS.findall(jd_text)
                        job["relevance_hits"] = len(hits)
                        print(f"  [{i+1}/{len(jobs)}] ✓ {job['title']} @ {job['company']} (relevance: {len(hits)})")
                        all_jobs.append(job)
                    if skipped:
                        print(f"  ↩ Skipped {skipped} already-seen jobs")
                    save_seen_ids(seen_ids)  # persist after each search term
                except Exception as e:
                    print(f"  ⚠ Error searching '{term}': {e}")
                    print(f"  Skipping to next search term...")
                    save_seen_ids(seen_ids)
                    continue

        # Deduplicate by job_id AND by title+company combo
        seen = set()
        seen_title_company = set()
        unique_jobs = []
        for j in all_jobs:
            key_tc = (j["title"].strip().lower(), j["company"].strip().lower())
            if j["job_id"] not in seen and key_tc not in seen_title_company:
                seen.add(j["job_id"])
                seen_title_company.add(key_tc)
                unique_jobs.append(j)

        print(f"\n{'='*60}")
        print(f"Total unique jobs collected: {len(unique_jobs)}")

        os.makedirs(LIST_DIR, exist_ok=True)
        output_path = os.path.join(LIST_DIR, OUTPUT_FILE)

        if ENABLE_SCORING:
            print("Scoring with Claude...")
            for i, job in enumerate(unique_jobs):
                if not job["description"]:
                    job["relevance_score"] = 0
                    job["relevance_reason"] = "Skipped (title filter)"
                    job["match_tags"] = ""
                    continue
                print(f"  Scoring [{i+1}/{len(unique_jobs)}]: {job['title']} @ {job['company']}")
                score, reason, tags = score_job(job["title"], job["company"], job["description"])
                job["relevance_score"] = score
                job["relevance_reason"] = reason
                job["match_tags"] = tags
                # Incremental save after each scoring
                _save_excel(unique_jobs[:i+1], output_path)
            unique_jobs.sort(key=lambda x: -x["relevance_score"])
        else:
            print("Scoring disabled (ENABLE_SCORING=False), skipping Claude API calls.")
            # Sort by relevance keyword hits (descending)
            unique_jobs.sort(key=lambda x: -x.get("relevance_hits", 0))
            for job in unique_jobs:
                job["relevance_score"] = job.get("relevance_hits", 0)
                job["relevance_reason"] = ""
                job["match_tags"] = ""

        # Build DataFrame
        df = pd.DataFrame(unique_jobs, columns=[
            "relevance_score", "relevance_reason", "match_tags",
            "title", "company", "location", "date_text",
            "easy_apply", "search_term", "url", "description"
        ])
        df.columns = [
            "相关性评分", "评分理由", "匹配关键词",
            "职位名称", "公司", "地点", "发布时间",
            "Easy Apply", "搜索词", "链接", "职位描述"
        ]

        # Write to Excel with formatting
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="职位列表")
            ws = writer.sheets["职位列表"]

            # Column widths
            ws.column_dimensions["A"].width = 10  # 评分
            ws.column_dimensions["B"].width = 40  # 评分理由
            ws.column_dimensions["C"].width = 30  # 关键词
            ws.column_dimensions["D"].width = 35  # 职位名称
            ws.column_dimensions["E"].width = 25  # 公司
            ws.column_dimensions["F"].width = 20  # 地点
            ws.column_dimensions["G"].width = 15  # 时间
            ws.column_dimensions["H"].width = 12  # Easy Apply
            ws.column_dimensions["I"].width = 25  # 搜索词
            ws.column_dimensions["J"].width = 50  # 链接
            ws.column_dimensions["K"].width = 60  # 描述

            # Color rows by score
            from openpyxl.styles import PatternFill, Font
            green  = PatternFill("solid", fgColor="C6EFCE")
            yellow = PatternFill("solid", fgColor="FFEB9C")
            red    = PatternFill("solid", fgColor="FFC7CE")

            if ENABLE_SCORING:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    score_val = row[0].value
                    if not isinstance(score_val, (int, float)):
                        continue
                    if score_val >= 7:
                        fill = green
                    elif score_val >= 5:
                        fill = yellow
                    else:
                        fill = red
                    for cell in row:
                        cell.fill = fill

        save_seen_ids(seen_ids)

        # Write to SQLite
        try:
            import db as _db
            _db.upsert_jobs(unique_jobs)
            print(f"   SQLite updated ({len(unique_jobs)} jobs)")
        except Exception as _e:
            print(f"   SQLite write failed: {_e}")

        print(f"\n✅ Done! Results saved to: {output_path}")
        print(f"   {len(unique_jobs)} new jobs collected")
        if ENABLE_SCORING:
            print(f"   {len(df[df['相关性评分'] >= 7])} high-relevance jobs (score 7+)")
            print(f"   {len(df[df['相关性评分'] >= 5])} medium-relevance jobs (score 5-6)")
            print(f"   {len(df[df['相关性评分'] < 5])} low-relevance jobs (score <5)")

    finally:
        driver.quit()

if __name__ == "__main__":
    main()
